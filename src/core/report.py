"""Báo cáo sau mỗi batch (CSV/Excel) và số liệu cho dashboard 30 ngày.

Báo cáo là thứ người dùng gửi cho kế toán hoặc lưu hồ sơ, nên phải đọc được bằng Excel
tiếng Việt: CSV ghi kèm BOM để Excel không hiển thị chữ có dấu thành rác.
"""

from __future__ import annotations

import csv
import logging
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from .learning import LearningStore
from .models import FileJob

logger = logging.getLogger(__name__)

COLUMNS = [
    "Tên cũ",
    "Tên mới",
    "Thư mục đích",
    "Profile",
    "Trạng thái",
    "Field trích được",
    "Tầng đã dùng",
    "Ghi chú",
    "Thời gian (ms)",
    "Mã băm nội dung",
]


def job_row(job: FileJob) -> list[str]:
    """1 dòng báo cáo cho 1 file."""
    return [
        job.source.name,
        job.new_name or "",
        str(job.dest_dir) if job.dest_dir else "",
        job.profile_name or "",
        job.status.label_vi,
        "; ".join(f"{k}={v.value}" for k, v in sorted(job.fields.items())),
        ", ".join(x.label_vi for x in job.layers_used),
        "; ".join(x for x in [job.message, *job.warnings] if x),
        str(job.duration_ms),
        job.file_hash[:16],
    ]


def default_report_name(prefix: str = "bao-cao", extension: str = ".csv") -> str:
    return f"{prefix}-{datetime.now().strftime('%Y%m%d-%H%M%S')}{extension}"


def write_csv(path: Path | str, jobs: list[FileJob]) -> Path:
    """Ghi CSV kèm BOM UTF-8 để Excel mở ra không bị lỗi font tiếng Việt."""
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(COLUMNS)
        for job in jobs:
            writer.writerow(job_row(job))
    return p


def write_excel(path: Path | str, jobs: list[FileJob]) -> Path:
    """Ghi .xlsx với dòng tiêu đề khóa sẵn và độ rộng cột vừa mắt."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Bao cao"

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for job in jobs:
        ws.append(job_row(job))

    widths = [30, 42, 34, 16, 12, 44, 26, 40, 14, 20]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = width
    ws.freeze_panes = "A2"

    wb.save(str(p))
    return p


def write_report(path: Path | str, jobs: list[FileJob]) -> Path:
    """Chọn định dạng theo đuôi file."""
    return write_excel(path, jobs) if str(path).lower().endswith(".xlsx") else write_csv(path, jobs)


# ------------------------------------------------------------------ dashboard


@dataclass
class ProfileStat:
    profile_id: str
    profile_name: str
    total: int
    success: int
    errors: int
    duplicates: int

    @property
    def success_rate(self) -> float:
        return round(self.success * 100.0 / self.total, 1) if self.total else 0.0

    @property
    def health(self) -> str:
        """Nhận xét ngắn để người dùng biết rule nào cần chỉnh."""
        if not self.total:
            return "Chưa có dữ liệu"
        if self.success_rate >= 95:
            return "Tốt"
        if self.success_rate >= 80:
            return "Cần để mắt"
        return "Nên chỉnh rule"


def profile_stats(
    learning: LearningStore, profile_names: dict[str, str], days: int = 30
) -> list[ProfileStat]:
    """Gộp số liệu match theo profile để hiển thị trong dashboard."""
    rows = learning.db.query(
        """
        SELECT profile_id,
               COUNT(*) AS total,
               SUM(CASE WHEN status = 'success' THEN 1 ELSE 0 END) AS success,
               SUM(CASE WHEN status = 'error' THEN 1 ELSE 0 END) AS errors,
               SUM(CASE WHEN status = 'duplicate' THEN 1 ELSE 0 END) AS duplicates
        FROM match_events
        WHERE ts >= datetime('now', ?)
        GROUP BY profile_id
        """,
        (f"-{int(days)} days",),
    )
    stats = [
        ProfileStat(
            profile_id=row["profile_id"],
            profile_name=profile_names.get(row["profile_id"], row["profile_id"] or "(không rõ)"),
            total=int(row["total"] or 0),
            success=int(row["success"] or 0),
            errors=int(row["errors"] or 0),
            duplicates=int(row["duplicates"] or 0),
        )
        for row in rows
    ]
    return sorted(stats, key=lambda s: (-s.total, s.profile_name))
