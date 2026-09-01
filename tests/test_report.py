"""Test báo cáo CSV/Excel và số liệu dashboard 30 ngày."""

from __future__ import annotations

import csv

from src.core.learning import LearningStore
from src.core.models import ExtractedField, FileJob, JobStatus, Layer
from src.core.report import (
    COLUMNS,
    ProfileStat,
    default_report_name,
    job_row,
    profile_stats,
    write_csv,
    write_excel,
    write_report,
)


def make_jobs(tmp_path) -> list[FileJob]:
    source = tmp_path / "chung-tu.pdf"
    source.write_bytes(b"%PDF")

    ok = FileJob(source=source, profile_id="inv", profile_name="Invoice")
    ok.status = JobStatus.SUCCESS
    ok.new_name = "2026-03-15_INV_A1.pdf"
    ok.dest_dir = tmp_path / "out" / "2026-08-31"
    ok.file_hash = "a" * 64
    ok.duration_ms = 412
    ok.layers_used = [Layer.TEXT, Layer.REGEX]
    ok.fields = {"number": ExtractedField(name="number", value="A1", layer=Layer.REGEX)}

    bad = FileJob(source=tmp_path / "hong.pdf", profile_name="")
    bad.status = JobStatus.ERROR
    bad.message = "PDF có mật khẩu"
    bad.warnings = ["Trùng số chứng từ"]
    return [ok, bad]


class TestJobRow:
    def test_du_cot(self, tmp_path):
        assert len(job_row(make_jobs(tmp_path)[0])) == len(COLUMNS)

    def test_noi_dung_dong(self, tmp_path):
        row = job_row(make_jobs(tmp_path)[0])
        assert row[0] == "chung-tu.pdf"
        assert row[1] == "2026-03-15_INV_A1.pdf"
        assert row[3] == "Invoice"
        assert row[4] == "Thành công"
        assert "number=A1" in row[5]
        assert "Regex theo nhãn" in row[6]
        assert row[8] == "412"

    def test_gop_ca_message_lan_canh_bao(self, tmp_path):
        row = job_row(make_jobs(tmp_path)[1])
        assert "PDF có mật khẩu" in row[7] and "Trùng số chứng từ" in row[7]


class TestCsv:
    def test_ghi_kem_bom_de_excel_khong_loi_font(self, tmp_path):
        path = write_csv(tmp_path / "bao-cao.csv", make_jobs(tmp_path))
        assert path.read_bytes().startswith(b"\xef\xbb\xbf")

    def test_doc_lai_dung_so_dong(self, tmp_path):
        path = write_csv(tmp_path / "bao-cao.csv", make_jobs(tmp_path))
        with path.open(encoding="utf-8-sig", newline="") as fh:
            rows = list(csv.reader(fh))
        assert rows[0] == COLUMNS
        assert len(rows) == 3  # tiêu đề + 2 file

    def test_giu_dau_tieng_viet(self, tmp_path):
        path = write_csv(tmp_path / "bao-cao.csv", make_jobs(tmp_path))
        assert "Thành công" in path.read_text(encoding="utf-8-sig")

    def test_batch_rong_van_ghi_tieu_de(self, tmp_path):
        path = write_csv(tmp_path / "rong.csv", [])
        assert "Tên cũ" in path.read_text(encoding="utf-8-sig")


class TestExcel:
    def test_ghi_va_doc_lai(self, tmp_path):
        from openpyxl import load_workbook

        path = write_excel(tmp_path / "bao-cao.xlsx", make_jobs(tmp_path))
        wb = load_workbook(path)
        ws = wb.active
        assert [c.value for c in ws[1]] == COLUMNS
        assert ws.cell(row=2, column=1).value == "chung-tu.pdf"
        assert ws.freeze_panes == "A2"
        wb.close()

    def test_chon_dinh_dang_theo_duoi_file(self, tmp_path):
        jobs = make_jobs(tmp_path)
        assert write_report(tmp_path / "a.xlsx", jobs).suffix == ".xlsx"
        assert write_report(tmp_path / "a.csv", jobs).suffix == ".csv"

    def test_ten_bao_cao_mac_dinh_co_thoi_diem(self):
        name = default_report_name()
        assert name.startswith("bao-cao-") and name.endswith(".csv")


class TestProfileStat:
    def test_ti_le_va_nhan_xet(self):
        assert ProfileStat("i", "Invoice", 100, 100, 0, 0).health == "Tốt"
        assert ProfileStat("i", "Invoice", 100, 90, 10, 0).health == "Cần để mắt"
        assert ProfileStat("i", "Invoice", 100, 50, 50, 0).health == "Nên chỉnh rule"
        assert ProfileStat("i", "Invoice", 0, 0, 0, 0).health == "Chưa có dữ liệu"

    def test_ti_le_thanh_cong(self):
        assert ProfileStat("i", "Invoice", 4, 3, 1, 0).success_rate == 75.0
        assert ProfileStat("i", "Invoice", 0, 0, 0, 0).success_rate == 0.0


class TestDashboard:
    def test_gop_theo_profile(self, db):
        learning = LearningStore(db)
        for _ in range(3):
            learning.record_match("inv", "success", "a.pdf")
        learning.record_match("inv", "error", "b.pdf")
        learning.record_match("inv", "duplicate", "c.pdf")
        learning.record_match("bl", "success", "d.pdf")

        stats = {s.profile_id: s for s in profile_stats(learning, {"inv": "Invoice"})}
        assert stats["inv"].profile_name == "Invoice"
        assert stats["inv"].total == 5
        assert stats["inv"].success == 3
        assert stats["inv"].errors == 1
        assert stats["inv"].duplicates == 1
        assert stats["bl"].profile_name == "bl"  # không có tên thì hiện id

    def test_sap_xep_theo_so_luong(self, db):
        learning = LearningStore(db)
        learning.record_match("it", "success")
        for _ in range(5):
            learning.record_match("nhieu", "success")
        assert profile_stats(learning, {})[0].profile_id == "nhieu"

    def test_khong_co_du_lieu(self, db):
        assert profile_stats(LearningStore(db), {}) == []
