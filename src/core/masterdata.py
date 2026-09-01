"""Tra cứu master data từ file Excel (vd mã KH -> tên công ty đầy đủ).

Load vào memory ở đầu batch, cache theo mtime, mở read-only để không đụng file
người dùng đang mở bằng Excel. File khóa/thiếu -> ném MasterDataError, pipeline bắt
lại thành cảnh báo trong Preview chứ KHÔNG làm hỏng cả batch (quyết định #8).
"""

from __future__ import annotations

import logging
import threading
from pathlib import Path

from .errors import MasterDataError
from .models import MasterDataLookup

logger = logging.getLogger(__name__)


def _column_index(header: list[str], spec: str) -> int:
    """Xác định cột theo tên tiêu đề (ưu tiên) hoặc theo chữ cái cột kiểu 'A', 'B'."""
    spec = (spec or "").strip()
    if not spec:
        raise MasterDataError("Chưa khai báo cột tra cứu trong profile")

    target = spec.casefold()
    for i, name in enumerate(header):
        if (name or "").strip().casefold() == target:
            return i

    if spec.isalpha() and len(spec) <= 3:
        index = 0
        for ch in spec.upper():
            index = index * 26 + (ord(ch) - ord("A") + 1)
        return index - 1

    raise MasterDataError(f"Không tìm thấy cột '{spec}' trong file master data")


class MasterDataTable:
    """1 bảng tra cứu đã nạp sẵn: key (đã casefold) -> value."""

    def __init__(
        self,
        mapping: dict[str, str],
        source: Path,
        mtime: float,
        originals: dict[str, str] | None = None,
    ) -> None:
        self.mapping = mapping
        # Khóa gốc như trong Excel — mapping lưu bản casefold nên không hiển thị được
        self.originals = originals or {}
        self.source = source
        self.mtime = mtime

    def lookup(self, key: str) -> str:
        return self.mapping.get((key or "").strip().casefold(), "")

    def first_pair(self) -> tuple[str, str] | None:
        """1 cặp key/value để làm ví dụ, key giữ đúng hoa/thường như trong file Excel."""
        for folded, value in self.mapping.items():
            return self.originals.get(folded, folded), value
        return None

    def __len__(self) -> int:
        return len(self.mapping)


def load_table(
    path: Path | str, key_column: str, value_column: str, sheet: str = ""
) -> MasterDataTable:
    """Đọc 1 sheet Excel thành bảng tra cứu. Dòng 1 là tiêu đề."""
    from openpyxl import load_workbook

    p = Path(path)
    if not p.exists():
        raise MasterDataError(f"Không tìm thấy file master data: {p}")

    try:
        wb = load_workbook(filename=str(p), read_only=True, data_only=True)
    except Exception as exc:
        raise MasterDataError(f"Không mở được file master data ({exc})") from exc

    try:
        if sheet:
            if sheet not in wb.sheetnames:
                raise MasterDataError(f"Không có sheet '{sheet}' trong {p.name}")
            ws = wb[sheet]
        else:
            ws = wb[wb.sheetnames[0]]

        rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            raise MasterDataError(f"Sheet rỗng: {p.name}") from None

        header = [str(c).strip() if c is not None else "" for c in header_row]
        ki = _column_index(header, key_column)
        vi = _column_index(header, value_column)

        mapping: dict[str, str] = {}
        originals: dict[str, str] = {}
        for row in rows:
            if row is None or ki >= len(row) or vi >= len(row):
                continue
            key, value = row[ki], row[vi]
            if key is None or value is None:
                continue
            k = str(key).strip()
            if k:
                # Dòng trùng key: giữ dòng đầu tiên cho deterministic
                mapping.setdefault(k.casefold(), str(value).strip())
                originals.setdefault(k.casefold(), k)
    finally:
        wb.close()

    return MasterDataTable(mapping, p, p.stat().st_mtime, originals)


class MasterDataStore:
    """Cache nhiều bảng theo (file, sheet, cột). Tự nạp lại khi mtime của file đổi."""

    def __init__(self, default_source: str = "") -> None:
        self.default_source = default_source
        self._cache: dict[tuple[str, str, str, str], MasterDataTable] = {}
        self._lock = threading.Lock()

    def resolve_source(self, spec: MasterDataLookup) -> Path:
        source = spec.source or self.default_source
        if not source:
            raise MasterDataError("Chưa cấu hình file Excel master data trong Settings")
        return Path(source)

    def get_table(self, spec: MasterDataLookup) -> MasterDataTable:
        path = self.resolve_source(spec)
        key = (str(path), spec.sheet, spec.key_column, spec.value_column)
        with self._lock:
            cached = self._cache.get(key)
            if cached is not None:
                try:
                    if cached.mtime == path.stat().st_mtime:
                        return cached
                except OSError:
                    pass  # file biến mất giữa chừng -> nạp lại để báo lỗi rõ ràng
            table = load_table(path, spec.key_column, spec.value_column, spec.sheet)
            self._cache[key] = table
            return table

    def lookup(self, spec: MasterDataLookup, key: str) -> str:
        """Tra 1 giá trị. Ném MasterDataError nếu nguồn dữ liệu có vấn đề."""
        if not key:
            return ""
        return self.get_table(spec).lookup(key)

    def clear(self) -> None:
        with self._lock:
            self._cache.clear()
